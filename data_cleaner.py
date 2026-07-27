"""
DatasetCleaner – automatic, non-blocking data cleaning for the StealthMonitor CSV dataset.

Design principles
─────────────────
• NEVER deletes any row, column, or attribute from the original file.
• NEVER renames, reorders, or drops columns from the original file.
• Produces a *cleaned copy* alongside the original file.
• Original file is ALWAYS preserved — only *_cleaned.csv is written to.
• Runs in a background thread so packet monitoring is never interrupted.
• Thread-safe: only one cleaning cycle runs at a time (re-entrant guard).
• Logs all cleaning actions via structured app_logger.
"""

from app_logger import get_logger
log = get_logger(__name__)

import csv
import os
import re
import threading
import time
import datetime
import unicodedata


# ──────────────────────────────────────────────────────────────────────
# Column classification (derived from the actual CSV schema)
# ──────────────────────────────────────────────────────────────────────

TIMESTAMP_COLUMNS = {
    "Timestamp", "Detection_Time", "Flow_Start_Time", "Last_Packet_Time",
}

BOOLEAN_COLUMNS = {
    "Encoded_Content", "Base64_Detected", "Shell_Command_Detected",
    "Exec_Command", "Eval_Command", "CMD_Usage", "Linux_Shell_Usage",
    "Passwd_Access_Attempt", "Admin_Mode", "Real_Block_Applied",
    "Simulated_Block",
}

# Columns that should carry integer flag values (0/1)
FLAG_COLUMNS = {
    "Reconnaissance_Flag", "Port_Scan_Flag", "Slow_Scan_Flag",
    "Multi_Port_Access", "Night_Activity", "Suspicious_Behavior",
    "Automated_Behavior", "DNS_Request", "HTTP_Request", "HTTPS_Request",
    "SSH_Attempt", "RDP_Attempt", "Database_Access", "Web_Attack_Indicator",
    "Directory_Traversal", "Remote_Code_Execution", "Exploit_Pattern",
    "Label", "ML_Prediction",
}

# Strictly numeric columns (int or float)
NUMERIC_COLUMNS = {
    "Source_Port", "Destination_Port", "Packet_Size", "Packet_Count",
    "Duration", "Rate", "Unique_Ports", "Connection_Count",
    "SYN_Count", "ACK_Count", "FIN_Count", "RST_Count", "PSH_Count",
    "URG_Count", "TTL_Value", "Window_Size", "Retransmission_Count",
    "Fragment_Count", "Payload_Size", "Payload_Entropy", "Payload_Length",
    "Keyword_Count", "DPI_Confidence", "AI_Confidence", "Threat_Score",
    "Behavioral_Score", "Anomaly_Score", "Stealth_Score",
    "Attack_Frequency", "Human_Behavior_Score", "Connection_Interval",
    "Session_Duration", "Active_Flow_Count", "VLAN_ID",
    "Threat_Count", "Attack_Count", "Safe_Traffic_Count",
    "Detection_Accuracy", "False_Positive_Rate", "False_Negative_Rate",
    "Precision_Score", "Recall_Score", "F1_Score",
    "Dataset_Row_ID", "CSV_Log_Row", "Prediction_Probability",
    "CPU_Usage", "Memory_Usage",
}

# Confidence / percentage columns – must fall within 0-100
CONFIDENCE_COLUMNS = {
    "DPI_Confidence", "AI_Confidence", "Threat_Score",
    "Detection_Accuracy", "Precision_Score", "Recall_Score", "F1_Score",
}

PORT_COLUMNS = {"Source_Port", "Destination_Port"}

PACKET_COUNT_COLUMNS = {
    "Packet_Count", "SYN_Count", "ACK_Count", "FIN_Count", "RST_Count",
    "PSH_Count", "URG_Count", "Retransmission_Count", "Fragment_Count",
    "Keyword_Count", "Threat_Count", "Attack_Count", "Safe_Traffic_Count",
    "Active_Flow_Count",
}

IP_COLUMNS = {"Source_IP", "Destination_IP", "Blocked_IP"}

# Categorical columns with their canonical values (used for normalisation)
CATEGORICAL_MAP = {
    "Protocol":           {"TCP", "UDP", "ICMP"},
    "Payload_Type":       {"Malicious", "Normal"},
    "DPI_Threat_Level":   {"Critical", "High", "Medium", "Low", "None"},
    "AI_Label":           {"POSITIVE", "NEGATIVE"},
    "Sentiment_Label":    {"POSITIVE", "NEGATIVE"},
    "Heuristic_Result":   {"Threat", "Safe"},
    "Scan_Intensity":     {"High", "Medium", "Low"},
    "Traffic_Type":       {"Malicious", "Benign"},
    "Packet_Direction":   {"Inbound", "Outbound"},
    "Flow_State":         {"Active", "Terminated"},
    "Firewall_Status":    {"Blocked", "Not Blocked", "Block Failed"},
    "Firewall_Action":    {"Blocked", "Not Blocked", "Block Failed"},
    "Log_Status":         {"Logged"},
    "Dashboard_Event":    {"Pushed"},
    "Dashboard_Connection": {"Active"},
    "Live_Alert":         {"Yes", "No"},
    "Security_Report_Status": {"Monitoring", "Generated"},
    "Model_Training_Status": {"Trained", "Untrained"},
    "RandomForest_Result": {"Malicious", "Normal"},
    "Dataset_Source":     {"Simulation", "Live", "Hybrid"},
    "System_Status":      {"Running", "Stopped", "Error"},
}

# Required fields that must not be empty after cleaning
REQUIRED_FIELDS = {
    "Timestamp", "Source_IP", "Destination_IP", "Protocol", "Event_ID",
}

# Regex for a valid IPv4 address
_IPV4_RE = re.compile(
    r"^(?:(?:25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(?:25[0-5]|2[0-4]\d|[01]?\d\d?)$"
)

# Non-printable character filter (keeps printable ASCII + common unicode)
_NON_PRINTABLE_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]")


# ──────────────────────────────────────────────────────────────────────
# Helper utilities
# ──────────────────────────────────────────────────────────────────────

def _is_null_like(value: str) -> bool:
    """Return True for values that represent missing / invalid data."""
    if value is None:
        return True
    v = str(value).strip().lower()
    return v in {"", "nan", "none", "null", "n/a", "na", "-", "undefined"}


def _safe_float(value, default=0.0):
    """Convert to float; return *default* on failure."""
    try:
        v = str(value).strip().replace("%", "").replace(",", "")
        return float(v)
    except (ValueError, TypeError):
        return default


def _safe_int(value, default=0):
    """Convert to int via float (handles '3.0' etc.); return *default* on failure."""
    try:
        return int(float(str(value).strip().replace(",", "")))
    except (ValueError, TypeError):
        return default


def _clean_string(value: str) -> str:
    """Strip whitespace and remove non-printable / corrupted characters."""
    if value is None:
        return ""
    s = str(value).strip()
    # Replace non-printable characters
    s = _NON_PRINTABLE_RE.sub("", s)
    # Normalise unicode (NFC form)
    s = unicodedata.normalize("NFC", s)
    return s


def _normalise_boolean(value: str) -> str:
    """Map assorted boolean representations to 'Yes'/'No' or 'True'/'False'."""
    v = str(value).strip().lower()
    if v in {"yes", "true", "1", "y"}:
        return "Yes"
    if v in {"no", "false", "0", "n", ""}:
        return "No"
    # For Admin_Mode, Real_Block_Applied, Simulated_Block the existing
    # format is Python's str(bool) → "True" / "False"
    return value.strip()


def _normalise_boolean_tf(value: str) -> str:
    """Normalise to 'True' / 'False' for Python-bool columns."""
    v = str(value).strip().lower()
    if v in {"true", "yes", "1", "y"}:
        return "True"
    if v in {"false", "no", "0", "n", ""}:
        return "False"
    return value.strip()


def _validate_ip(value: str) -> str:
    """Return the IP if it's a valid IPv4 address, otherwise empty string."""
    v = str(value).strip()
    if not v or _is_null_like(v):
        return ""
    if _IPV4_RE.match(v):
        return v
    return ""


def _normalise_timestamp(value: str) -> str:
    """Try to parse and re-format timestamps into a canonical form."""
    v = str(value).strip()
    if _is_null_like(v):
        return ""
    # Already in our canonical format?
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y %H:%M:%S",
                "%m/%d/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f"):
        try:
            dt = datetime.datetime.strptime(v, fmt)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    # Return as-is if we can't parse – never delete data
    return v


def _normalise_categorical(value: str, canonical_set: set) -> str:
    """Match value to the closest canonical category (case-insensitive)."""
    v = _clean_string(value)
    if _is_null_like(v):
        return ""
    lookup = {c.lower(): c for c in canonical_set}
    vl = v.lower()
    if vl in lookup:
        return lookup[vl]
    # No match – preserve original (never discard)
    return v


# ──────────────────────────────────────────────────────────────────────
# Core cleaner
# ──────────────────────────────────────────────────────────────────────

class DatasetCleaner:
    """
    Reads the raw CSV, cleans every field, de-duplicates exact rows, and
    writes a *_cleaned.csv alongside the original.  The original is
    NEVER modified.
    """

    def __init__(self, source_csv: str, *, rows_between_cleans: int = 50):
        self.source_csv = os.path.abspath(source_csv)
        self._base, self._ext = os.path.splitext(self.source_csv)
        self.cleaned_csv = self._base + "_cleaned" + self._ext
        self.cleaned_xlsx = self._base + "_cleaned.xlsx"

        # Trigger cleaning every N new rows (configurable)
        self._rows_between_cleans = rows_between_cleans
        self._rows_since_clean = 0
        self._lock = threading.Lock()
        self._cleaning_in_progress = False

    def notify_new_row(self):
        with self._lock:
            self._rows_since_clean += 1
            if self._rows_since_clean >= self._rows_between_cleans:
                self._rows_since_clean = 0
                if not self._cleaning_in_progress:
                    self._cleaning_in_progress = True
                    threading.Thread(target=self._run_cleaning_cycle, daemon=True).start()


        # Latest statistics
        self.last_stats: dict = {}
        
        # Incremental state
        self._last_position = 0
        self._columns = []
        self._seen_event_ids = set()
        self._row_fingerprints = set()

    # ── public API ────────────────────────────────────────────────────

    def notify_new_row(self):
        """Called by the logger after every successful row append."""
        with self._lock:
            self._rows_since_clean += 1
            if self._rows_since_clean >= self._rows_between_cleans and not self._cleaning_in_progress:
                self._rows_since_clean = 0
                self._cleaning_in_progress = True
                t = threading.Thread(target=self._run_clean, daemon=True)
                t.start()

    def force_clean(self):
        """Trigger a cleaning cycle immediately (non-blocking)."""
        with self._lock:
            if self._cleaning_in_progress:
                return
            self._cleaning_in_progress = True
            self._rows_since_clean = 0
        t = threading.Thread(target=self._run_clean, daemon=True)
        t.start()

    # ── internal ──────────────────────────────────────────────────────

    def _run_clean(self):
        """Execute an incremental cleaning pass in the background with file locking.

        IMPORTANT: This method NEVER modifies the original source CSV file.
        It only appends cleaned rows to the _cleaned.csv sibling file.
        """
        stats = {
            "total_rows": 0,
            "cleaned_rows": 0,
            "duplicate_rows_removed": 0,
            "missing_values_corrected": 0,
            "invalid_values_corrected": 0,
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        lock_file = self.source_csv + ".lock"
        try:
            # File lock to prevent concurrent cross-process cleaning
            fd = os.open(lock_file, os.O_CREAT | os.O_EXCL | os.O_RDWR)
            os.close(fd)
        except FileExistsError:
            with self._lock:
                self._cleaning_in_progress = False
            return

        try:
            if not os.path.exists(self.source_csv):
                return

            # ── 1. Incremental Read ───────────────────────────────────
            rows = []
            try:
                with open(self.source_csv, newline="", encoding="utf-8") as fh:
                    fh.seek(self._last_position)
                    if self._last_position == 0:
                        reader = csv.DictReader(fh)
                        if reader.fieldnames is None:
                            return
                        self._columns = list(reader.fieldnames)
                    else:
                        reader = csv.DictReader(fh, fieldnames=self._columns)
                    
                    for row in reader:
                        rows.append(row)
                    self._last_position = fh.tell()
            except Exception as e:
                print(f"[DataCleaner] Read error: {e}")
                return

            stats["total_rows"] = len(rows)
            if not rows:
                self._write_stats(stats)
                return

            # ── 2. Clean every cell ───────────────────────────────────
            cleaned_rows = []

            for row in rows:
                dirty = False  # track if anything was changed

                for col in self._columns:
                    original_val = row.get(col, "")
                    new_val = self._clean_cell(col, original_val, stats)
                    if new_val != original_val:
                        dirty = True
                    row[col] = new_val

                # ── Event_ID uniqueness ───────────────────────────────
                eid = row.get("Event_ID", "")
                if eid and eid in self._seen_event_ids:
                    # Make unique by appending suffix
                    suffix = 1
                    while f"{eid}_dup{suffix}" in self._seen_event_ids:
                        suffix += 1
                    row["Event_ID"] = f"{eid}_dup{suffix}"
                    stats["invalid_values_corrected"] += 1
                    dirty = True
                if row.get("Event_ID"):
                    self._seen_event_ids.add(row["Event_ID"])

                # ── Exact-duplicate removal (fingerprint) ─────────────
                fp = tuple(row.get(c, "") for c in self._columns)
                if fp in self._row_fingerprints:
                    stats["duplicate_rows_removed"] += 1
                    continue  # skip duplicate
                self._row_fingerprints.add(fp)

                if dirty:
                    stats["cleaned_rows"] += 1

                cleaned_rows.append(row)

            # ── 3. Append to cleaned CSV ──────────────────────────────
            if not cleaned_rows:
                self._write_stats(stats)
                return
                
            try:
                write_header = not os.path.exists(self.cleaned_csv)
                with open(self.cleaned_csv, mode="a", newline="", encoding="utf-8") as fh:
                    writer = csv.DictWriter(fh, fieldnames=self._columns)
                    if write_header:
                        writer.writeheader()
                    writer.writerows(cleaned_rows)
            except Exception as e:
                print(f"[DataCleaner] Write CSV error: {e}")
                return

            # ── 4. Log stats ──────────────────────────────────────────
            self._write_stats(stats)
            self.last_stats = stats

            log.info(
                "Incremental clean complete",
                extra={
                    "processed": stats['total_rows'],
                    "cleaned": stats['cleaned_rows'],
                    "duplicates_removed": stats['duplicate_rows_removed'],
                    "missing_corrected": stats['missing_values_corrected'],
                    "invalid_corrected": stats['invalid_values_corrected'],
                    "source_file_modified": False,  # Explicit guarantee
                }
            )

        except Exception as e:
            print(f"[DataCleaner] Unexpected error: {e}")
        finally:
            if os.path.exists(lock_file):
                try:
                    os.remove(lock_file)
                except OSError:
                    pass
            with self._lock:
                self._cleaning_in_progress = False

    # ── Per-cell cleaning logic ───────────────────────────────────────

    def _clean_cell(self, col: str, value, stats: dict) -> str:
        """Clean a single cell. Returns the cleaned string value."""
        val = _clean_string(value)

        # ── Timestamps ────────────────────────────────────────────────
        if col in TIMESTAMP_COLUMNS:
            cleaned = _normalise_timestamp(val)
            if cleaned != val:
                stats["invalid_values_corrected"] += 1
            if _is_null_like(cleaned) and col in REQUIRED_FIELDS:
                cleaned = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                stats["missing_values_corrected"] += 1
            return cleaned

        # ── Boolean columns (Yes/No style) ────────────────────────────
        if col in BOOLEAN_COLUMNS:
            if _is_null_like(val):
                stats["missing_values_corrected"] += 1
                # Columns that use True/False
                if col in {"Admin_Mode", "Real_Block_Applied", "Simulated_Block"}:
                    return "False"
                return "No"
            if col in {"Admin_Mode", "Real_Block_Applied", "Simulated_Block"}:
                cleaned = _normalise_boolean_tf(val)
            else:
                cleaned = _normalise_boolean(val)
            if cleaned != val:
                stats["invalid_values_corrected"] += 1
            return cleaned

        # ── Flag columns (0/1 integers) ───────────────────────────────
        if col in FLAG_COLUMNS:
            if _is_null_like(val):
                stats["missing_values_corrected"] += 1
                return "0"
            iv = _safe_int(val, default=-1)
            if iv not in (0, 1):
                # Try boolean interpretation
                vl = val.lower()
                if vl in {"yes", "true", "y"}:
                    stats["invalid_values_corrected"] += 1
                    return "1"
                stats["invalid_values_corrected"] += 1
                return "0"
            s = str(iv)
            if s != val:
                stats["invalid_values_corrected"] += 1
            return s

        # ── IP columns ────────────────────────────────────────────────
        if col in IP_COLUMNS:
            if _is_null_like(val):
                # Blocked_IP can be legitimately empty
                if col == "Blocked_IP":
                    return ""
                stats["missing_values_corrected"] += 1
                return "0.0.0.0"
            cleaned = _validate_ip(val)
            if not cleaned and val:
                # Invalid format – preserve original, flag it
                stats["invalid_values_corrected"] += 1
                return val  # never delete
            return cleaned if cleaned else val

        # ── Port columns ──────────────────────────────────────────────
        if col in PORT_COLUMNS:
            if _is_null_like(val):
                stats["missing_values_corrected"] += 1
                return "0"
            iv = _safe_int(val, default=-1)
            if iv < 0 or iv > 65535:
                stats["invalid_values_corrected"] += 1
                return str(max(0, min(65535, iv))) if iv >= 0 else "0"
            s = str(iv)
            if s != val:
                stats["invalid_values_corrected"] += 1
            return s

        # ── Confidence / percentage columns (0-100) ───────────────────
        if col in CONFIDENCE_COLUMNS:
            if _is_null_like(val):
                stats["missing_values_corrected"] += 1
                return "0.0"
            fv = _safe_float(val, default=-1.0)
            if fv < 0:
                stats["invalid_values_corrected"] += 1
                return "0.0"
            if fv > 100:
                stats["invalid_values_corrected"] += 1
                fv = 100.0
            s = str(round(fv, 2))
            return s

        # ── Packet count / non-negative integer columns ───────────────
        if col in PACKET_COUNT_COLUMNS:
            if _is_null_like(val):
                stats["missing_values_corrected"] += 1
                return "0"
            iv = _safe_int(val, default=-1)
            if iv < 0:
                stats["invalid_values_corrected"] += 1
                return "0"
            s = str(iv)
            if s != val:
                stats["invalid_values_corrected"] += 1
            return s

        # ── Other numeric columns ─────────────────────────────────────
        if col in NUMERIC_COLUMNS:
            if _is_null_like(val):
                stats["missing_values_corrected"] += 1
                return "0"
            fv = _safe_float(val, default=None)
            if fv is None:
                stats["invalid_values_corrected"] += 1
                return "0"
            # Keep precision: use int if it looks integral
            if fv == int(fv) and "." not in val:
                return str(int(fv))
            return str(round(fv, 4))

        # ── Categorical columns ───────────────────────────────────────
        if col in CATEGORICAL_MAP:
            canonical = CATEGORICAL_MAP[col]
            if _is_null_like(val):
                stats["missing_values_corrected"] += 1
                # Provide a sensible default
                defaults = {
                    "Protocol": "TCP",
                    "Payload_Type": "Normal",
                    "DPI_Threat_Level": "None",
                    "AI_Label": "POSITIVE",
                    "Sentiment_Label": "POSITIVE",
                    "Heuristic_Result": "Safe",
                    "Scan_Intensity": "Low",
                    "Traffic_Type": "Benign",
                    "Packet_Direction": "Inbound",
                    "Flow_State": "Active",
                    "Firewall_Status": "Not Blocked",
                    "Firewall_Action": "Not Blocked",
                    "Log_Status": "Logged",
                    "Dashboard_Event": "Pushed",
                    "Dashboard_Connection": "Active",
                    "Live_Alert": "No",
                    "Security_Report_Status": "Monitoring",
                    "Model_Training_Status": "Trained",
                    "RandomForest_Result": "Normal",
                    "Dataset_Source": "Simulation",
                    "System_Status": "Running",
                }
                return defaults.get(col, "")
            cleaned = _normalise_categorical(val, canonical)
            if cleaned != val:
                stats["invalid_values_corrected"] += 1
            return cleaned

        # ── Required fields fallback ──────────────────────────────────
        if col in REQUIRED_FIELDS and _is_null_like(val):
            stats["missing_values_corrected"] += 1
            if col == "Event_ID":
                return f"EVT-RECOVERED-{int(time.time()*1000)}"
            return val  # keep whatever is there

        # ── Generic string cleanup ────────────────────────────────────
        if _is_null_like(val) and col not in {
            "Malicious_Keyword", "Block_Rule_Name", "Blocked_IP",
            "DPI_Result", "Heuristic_Trigger", "Threat_Intelligence",
            "GenAI_Result", "Transformer_Output", "Feature_Importance",
        }:
            # Only fill non-nullable string columns
            pass  # keep empty – these are legitimately empty for benign traffic

        return val

    # ── XLSX writer ───────────────────────────────────────────────────

    def _write_xlsx(self, columns: list, rows: list[dict]):
        """Write an XLSX copy if openpyxl is available. Gracefully skips otherwise."""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cleaned Security Logs"

            # Header row
            for c_idx, col_name in enumerate(columns, 1):
                ws.cell(row=1, column=c_idx, value=col_name)

            # Data rows
            for r_idx, row in enumerate(rows, 2):
                for c_idx, col_name in enumerate(columns, 1):
                    ws.cell(row=r_idx, column=c_idx, value=row.get(col_name, ""))

            wb.save(self.cleaned_xlsx)
        except ImportError:
            # openpyxl not installed – skip XLSX generation silently
            pass
        except Exception as e:
            print(f"[DataCleaner] XLSX write error (non-fatal): {e}")

    # ── Stats logging ─────────────────────────────────────────────────

    def _write_stats(self, stats: dict):
        """Append cleaning statistics to a log file."""
        stats_file = self._base + "_cleaning_stats.log"
        try:
            with open(stats_file, mode="a", encoding="utf-8") as fh:
                fh.write(
                    f"[{stats['timestamp']}] "
                    f"Total Rows: {stats['total_rows']} | "
                    f"Cleaned Rows: {stats['cleaned_rows']} | "
                    f"Duplicate Rows Removed: {stats['duplicate_rows_removed']} | "
                    f"Missing Values Corrected: {stats['missing_values_corrected']} | "
                    f"Invalid Values Corrected: {stats['invalid_values_corrected']} | "
                    f"Source File Modified: False\n"
                )
        except Exception as e:
            log.error(f"Stats write error: {e}")
